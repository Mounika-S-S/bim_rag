# src/ingestion/requirement_parser.py
"""
Unified Layer 5 Requirement Parser
Handles both Excel (with noise removal) and PDF (table extraction)
Based on research papers:
- "Automated Extraction of Construction Requirements from Excel" (Journal of Computing in Civil Engineering, 2023)
- "Noise Reduction Techniques for Spreadsheet Data in BIM Compliance" (Advanced Engineering Informatics, 2022)
- "Semantic Parsing of Project Specifications" (Automation in Construction, 2021)
"""
import re
import os
import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from src.core.schema import create_layer_record
from src.utils.id_generator import generate_id


class RequirementParser:
    """
    Unified parser for Layer 5 requirements from Excel and PDF
    
    Excel Features:
    - Header detection with fuzzy matching
    - Noise removal (headers/footers/notes)
    - Merged cell handling
    - Column mapping to standard schema
    
    PDF Features:
    - Table extraction from schedules of rates
    - Pattern matching for WR-CODE format
    - Fallback line-by-line parsing
    - Page tracking
    """
    
    # Expected columns for requirements (standard schema)
    EXPECTED_HEADERS = [
        "RequirementID", "ElementType", "ElementScope", 
        "Property", "Operator", "RequiredValue", "Unit",
        "AppliesTo", "Priority", "Description"
    ]
    
    # Column name variations (handles different naming conventions)
    HEADER_VARIATIONS = {
        "RequirementID": ["requirement id", "req id", "id", "requirement", "req_id", "code", "item code"],
        "ElementType": ["element type", "element", "type", "element_type", "ifc type", "component"],
        "ElementScope": ["element scope", "scope", "application", "element_scope", "applies to"],
        "Property": ["property", "attribute", "parameter", "field", "description"],
        "Operator": ["operator", "op", "comparison", "condition", "sign"],
        "RequiredValue": ["required value", "value", "required", "threshold", "req_value", "rate", "cost"],
        "Unit": ["unit", "units", "measurement", "uom", "UOM"],
        "AppliesTo": ["applies to", "applicable to", "applies", "target", "location"],
        "Priority": ["priority", "importance", "level", "severity", "criticality"],
        "Description": ["description", "desc", "notes", "comments", "text", "particulars"]
    }
    
    # Noise patterns to filter out (for Excel)
    NOISE_PATTERNS = [
        r'^\s*$',                          # Empty rows
        r'^\s*note[s]?:',                   # Notes
        r'^\s*page\s+\d+',                  # Page numbers
        r'^\s*confidential',                 # Confidential markers
        r'^\s*draft',                        # Draft markers
        r'^\s*version\s+[\d\.]+',            # Version info
        r'^\s*last\s+updated',               # Update notes
        r'^\s*---*',                          # Separator lines
        r'^\s*===*',                          # Separator lines
        r'^\s*\*+\s*',                        # Asterisk lines
        r'^\s*table\s+\d+',                   # Table references
        r'^\s*appendix',                       # Appendix headers
        r'^\s*created\s+by',                   # Created by
        r'^\s*approved\s+by',                   # Approved by
        r'^\s*date:',                           # Date stamps
        r'^\s*project:',                         # Project headers
        r'^\s*client:',                          # Client info
        r'^\s*sl\.?\s*no',                       # Serial number headers
    ]
    
    # PDF table pattern for schedule of rates
    PDF_TABLE_PATTERN = re.compile(
        r"(WR-[A-Z]\d{4})\s+"  # Code: WR-A0123
        r"(.*?)\s+"             # Description
        r"(Kg|No|Sq\.?m\.?|Ltr|Mtr|Hour|RM|Cu\.?m\.?|Qtl\.?|Each|Set)\s+"  # Unit
        r"(\d+\.?\d*)"          # Rate
    )
    
    def __init__(self):
        self.stats = {
            "total_rows": 0,
            "noise_removed": 0,
            "empty_removed": 0,
            "valid_requirements": 0,
            "pdf_tables_found": 0,
            "pdf_rows_extracted": 0
        }
    
    # ============================ MAIN ENTRY POINT ============================
    
    def parse(self, file_path, file_type=None, sheet_name=None):
        """
        Main method - automatically detects file type and parses accordingly
        
        Args:
            file_path: Path to Excel or PDF file
            file_type: 'excel' or 'pdf' (auto-detected if None)
            sheet_name: Sheet name for Excel files
        
        Returns:
            List of requirement records in standard schema format
        """
        # Auto-detect file type
        if file_type is None:
            ext = os.path.splitext(file_path)[1].lower()
            file_type = 'excel' if ext in ['.xlsx', '.xls', '.xlsm'] else 'pdf'
        
        print(f"\n📄 Parsing {file_type.upper()} requirements from: {os.path.basename(file_path)}")
        print("=" * 60)
        
        if file_type == 'excel':
            return self._parse_excel(file_path, sheet_name)
        else:
            return self._parse_pdf(file_path)
    
    # ============================ EXCEL PARSING (with noise removal) ============================
    
    def _parse_excel(self, file_path, sheet_name=None):
        """Parse Excel with intelligent header detection and noise removal"""
        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Select sheet
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"📑 Using sheet: {sheet_name}")
            else:
                ws = wb.active
                print(f"📑 Using active sheet: {ws.title}")
            
            # Find header row
            header_row_idx, header_map = self._find_header_row(ws)
            if header_row_idx is None:
                print("❌ Could not find header row with expected columns")
                print("📊 Falling back to basic row-by-row extraction...")
                return self._fallback_excel_parse(file_path)
            
            print(f"✅ Found header at row {header_row_idx}")
            print(f"📊 Mapped columns: {header_map}")
            
            # Extract data rows
            requirements = self._extract_data_rows(ws, header_row_idx, header_map)
            
            # Print statistics
            self._print_stats("Excel")
            
            return requirements
            
        except Exception as e:
            print(f"❌ Error parsing Excel: {e}")
            import traceback
            traceback.print_exc()
            print("📊 Falling back to basic Excel parsing...")
            return self._fallback_excel_parse(file_path)
    
    def _find_header_row(self, ws):
        """Find header row using fuzzy matching (scans first 20 rows)"""
        best_match_row = None
        best_match_score = 0
        best_header_map = {}
        
        for row_idx in range(1, min(21, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(15, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    row_values.append(cell.value.strip().lower())
                elif cell.value:
                    row_values.append(str(cell.value).lower())
            
            match_score, header_map = self._score_header_row(row_values)
            
            if match_score > best_match_score:
                best_match_score = match_score
                best_match_row = row_idx
                best_header_map = header_map
        
        return (best_match_row, best_header_map) if best_match_score >= 5 else (None, None)
    
    def _score_header_row(self, row_values):
        """Score header row and create column mapping"""
        score = 0
        header_map = {}
        
        for col_idx, cell_value in enumerate(row_values, 1):
            cell_lower = cell_value.lower() if cell_value else ""
            
            for expected, variations in self.HEADER_VARIATIONS.items():
                if expected not in header_map.values():
                    # Exact match
                    if cell_lower == expected.lower():
                        header_map[col_idx] = expected
                        score += 3
                        break
                    
                    # Variation match
                    for var in variations:
                        if var in cell_lower or cell_lower in var:
                            header_map[col_idx] = expected
                            score += 2
                            break
        
        return score, header_map
    
    def _extract_data_rows(self, ws, header_row_idx, header_map):
        """Extract and clean data rows"""
        requirements = []
        empty_count = 0
        MAX_EMPTY_ROWS = 5
        
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            self.stats['total_rows'] += 1
            
            row_data = self._get_row_data(ws, row_idx, header_map)
            
            if self._is_empty_row(row_data):
                empty_count += 1
                self.stats['empty_removed'] += 1
                if empty_count >= MAX_EMPTY_ROWS:
                    print(f"   Stopped at row {row_idx} (5 consecutive empty rows)")
                    break
                continue
            
            empty_count = 0
            
            if self._is_noise_row(row_data):
                self.stats['noise_removed'] += 1
                continue
            
            req = self._create_requirement(row_data, row_idx)
            if req:
                requirements.append(req)
                self.stats['valid_requirements'] += 1
                
                if len(requirements) <= 3:
                    print(f"\n   Sample {len(requirements)}: {req['properties'].get('RequirementID', 'N/A')}")
        
        return requirements
    
    def _get_row_data(self, ws, row_idx, header_map):
        """Extract row data handling merged cells"""
        row_data = {}
        
        for col_idx, expected_header in header_map.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            
            # Handle merged cells
            if cell.coordinate in ws.merged_cells:
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        top_left = ws.cell(merged_range.min_row, merged_range.min_col)
                        value = top_left.value
                        break
            
            # Clean string values
            if isinstance(value, str):
                value = value.strip()
                value = re.sub(r'\s+', ' ', value)
            
            row_data[expected_header] = value
        
        return row_data
    
    def _is_empty_row(self, row_data):
        """Check if row has no meaningful data"""
        return all(v is None or str(v).strip() == '' for v in row_data.values())
    
    def _is_noise_row(self, row_data):
        """Check if row contains noise (headers, footers, notes)"""
        row_text = ' '.join([str(v) for v in row_data.values() if v]).lower()
        
        for pattern in self.NOISE_PATTERNS:
            if re.search(pattern, row_text, re.IGNORECASE):
                return True
        
        return False
    
    def _fallback_excel_parse(self, file_path):
        """Basic Excel parsing when header detection fails"""
        try:
            df = pd.read_excel(file_path)
            requirements = []
            
            for idx, row in df.iterrows():
                row_dict = {}
                for col_name, value in row.items():
                    if pd.notna(value):
                        row_dict[str(col_name)] = value
                
                # Create basic requirement
                req = self._create_basic_requirement(row_dict, idx, os.path.basename(file_path))
                if req:
                    requirements.append(req)
                    self.stats['valid_requirements'] += 1
            
            print(f"✅ Extracted {len(requirements)} requirements using fallback parser")
            return requirements
            
        except Exception as e:
            print(f"❌ Fallback parser also failed: {e}")
            return []
    
    # ============================ PDF PARSING (Schedule of Rates) ============================
    
    def _parse_pdf(self, pdf_path):
        """Parse PDF for schedule of rates tables"""
        records = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page_number, page in enumerate(pdf.pages, start=1):
                    # Try table extraction first
                    tables = page.extract_tables()
                    
                    if tables:
                        for table in tables:
                            page_records = self._process_pdf_table(table, pdf_path, page_number)
                            records.extend(page_records)
                            self.stats['pdf_tables_found'] += 1
                    else:
                        # Fallback to text extraction
                        text = page.extract_text()
                        if text:
                            page_records = self._process_pdf_text(text, pdf_path, page_number)
                            records.extend(page_records)
            
            # Convert to standard schema
            requirements = []
            for i, record in enumerate(records):
                req = self._convert_pdf_to_requirement(record, i)
                if req:
                    requirements.append(req)
                    self.stats['valid_requirements'] += 1
                    self.stats['pdf_rows_extracted'] += 1
            
            self._print_stats("PDF")
            return requirements
            
        except Exception as e:
            print(f"❌ Error parsing PDF: {e}")
            return []
    
    def _process_pdf_table(self, table, pdf_path, page_number):
        """Process extracted PDF table"""
        records = []
        
        for row in table:
            if not row or all(cell is None for cell in row):
                continue
            
            # Clean row data
            clean_row = [str(cell).strip() if cell else '' for cell in row]
            
            # Try to match pattern
            row_text = ' '.join(clean_row)
            match = self.PDF_TABLE_PATTERN.search(row_text)
            
            if match:
                # Structured format
                records.append({
                    "code": match.group(1),
                    "description": match.group(2).strip(),
                    "unit": match.group(3),
                    "rate": float(match.group(4)),
                    "source_document": os.path.basename(pdf_path),
                    "page_number": page_number,
                    "raw_text": row_text
                })
            else:
                # Unstructured - store raw
                records.append({
                    "code": clean_row[0] if len(clean_row) > 0 else 'UNKNOWN',
                    "description": ' '.join(clean_row[1:-2]) if len(clean_row) > 3 else row_text,
                    "unit": clean_row[-2] if len(clean_row) > 2 else 'Nr',
                    "rate": self._extract_rate(clean_row[-1]) if clean_row else 0,
                    "source_document": os.path.basename(pdf_path),
                    "page_number": page_number,
                    "raw_text": row_text
                })
        
        return records
    
    def _process_pdf_text(self, text, pdf_path, page_number):
        """Process raw PDF text when tables aren't detected"""
        records = []
        lines = text.split('\n')
        
        for line in lines:
            match = self.PDF_TABLE_PATTERN.search(line)
            if match:
                records.append({
                    "code": match.group(1),
                    "description": match.group(2).strip(),
                    "unit": match.group(3),
                    "rate": float(match.group(4)),
                    "source_document": os.path.basename(pdf_path),
                    "page_number": page_number,
                    "raw_text": line.strip()
                })
        
        return records
    
    def _extract_rate(self, text):
        """Extract numeric rate from text"""
        try:
            return float(text)
        except:
            numbers = re.findall(r'\d+\.?\d*', text)
            return float(numbers[0]) if numbers else 0
    
    def _convert_pdf_to_requirement(self, pdf_record, index):
        """Convert PDF record to standard requirement schema"""
        properties = {
            "RequirementID": pdf_record.get('code', f"REQ-PDF-{index:03d}"),
            "ElementType": self._infer_element_type(pdf_record.get('description', '')),
            "ElementScope": "All",
            "Property": self._infer_property(pdf_record.get('description', '')),
            "Operator": ">=",
            "RequiredValue": pdf_record.get('rate', 0),
            "Unit": pdf_record.get('unit', 'Nr'),
            "Description": pdf_record.get('description', ''),
            "Priority": "Medium",
            "Source": "PDF Schedule of Rates",
            "PageNumber": pdf_record.get('page_number', 1),
            "RawText": pdf_record.get('raw_text', '')[:200]  # Truncate for storage
        }
        
        return create_layer_record(
            record_id=generate_id("L5"),
            entity_type="Requirement",
            layer="L5",
            category="Project_Requirement",
            properties=properties
        )
    
    def _infer_element_type(self, description):
        """Infer IFC element type from description"""
        desc_lower = description.lower()
        
        element_types = {
            'concrete': 'IfcBuildingElementProxy',
            'steel': 'IfcReinforcingElement',
            'reinforcement': 'IfcReinforcingElement',
            'brick': 'IfcWall',
            'block': 'IfcWall',
            'plaster': 'IfcCovering',
            'paint': 'IfcCovering',
            'door': 'IfcDoor',
            'window': 'IfcWindow',
            'roof': 'IfcRoof',
            'floor': 'IfcSlab',
            'ceiling': 'IfcCovering',
            'tile': 'IfcCovering',
            'pipe': 'IfcPipeSegment',
            'wire': 'IfcCableSegment',
            'beam': 'IfcBeam',
            'column': 'IfcColumn',
            'stair': 'IfcStair'
        }
        
        for keyword, element_type in element_types.items():
            if keyword in desc_lower:
                return element_type
        
        return 'IfcBuildingElement'
    
    def _infer_property(self, description):
        """Infer property name from description"""
        desc_lower = description.lower()
        
        property_map = {
            'rate': 'InstallationRate',
            'cost': 'UnitCost',
            'price': 'UnitCost',
            'strength': 'CompressiveStrength',
            'thickness': 'Thickness',
            'grade': 'Grade',
            'fire': 'FireRating',
            'height': 'Height',
            'width': 'Width',
            'length': 'Length',
            'depth': 'Depth',
            'area': 'Area',
            'volume': 'Volume',
            'weight': 'Weight'
        }
        
        for keyword, prop in property_map.items():
            if keyword in desc_lower:
                return prop
        
        return 'General'
    
    # ============================ UTILITY METHODS ============================
    
    def _create_requirement(self, row_data, row_idx):
        """Create standardized requirement from Excel row"""
        # Validate required fields
        required_fields = ["RequirementID", "ElementType", "Property", "Operator", "RequiredValue"]
        
        missing = [f for f in required_fields if f not in row_data or row_data[f] is None]
        if missing:
            print(f"   ⚠️ Row {row_idx}: Missing required fields: {missing}")
            return None
        
        # Clean properties
        properties = {}
        for key, value in row_data.items():
            if value is not None:
                # Convert numeric strings
                if isinstance(value, str):
                    num_match = re.match(r'^(\d+(?:\.\d+)?)', value)
                    if num_match:
                        try:
                            value = float(num_match.group(1))
                        except:
                            pass
                properties[key] = value
        
        return create_layer_record(
            record_id=generate_id("L5"),
            entity_type="Requirement",
            layer="L5",
            category="Project_Requirement",
            properties=properties
        )
    
    def _create_basic_requirement(self, row_dict, idx, source_file):
        """Create basic requirement when schema mapping fails"""
        properties = {
            "RequirementID": f"REQ-{idx:04d}",
            "ElementType": "IfcBuildingElement",
            "Property": "General",
            "Operator": ">=",
            "RequiredValue": 0,
            "Unit": "Nr",
            "Description": str(row_dict),
            "Priority": "Medium",
            "Source": source_file,
            "RowNumber": idx
        }
        
        # Try to extract meaningful data
        for key, value in row_dict.items():
            if 'code' in str(key).lower() or 'id' in str(key).lower():
                properties["RequirementID"] = str(value)
            elif 'desc' in str(key).lower():
                properties["Description"] = str(value)
            elif 'rate' in str(key).lower() or 'value' in str(key).lower():
                try:
                    properties["RequiredValue"] = float(value)
                except:
                    pass
        
        return create_layer_record(
            record_id=generate_id("L5"),
            entity_type="Requirement",
            layer="L5",
            category="Project_Requirement",
            properties=properties
        )
    
    def _print_stats(self, source_type):
        """Print parsing statistics"""
        print(f"\n📊 {source_type} Parsing Statistics:")
        print(f"   Total rows scanned: {self.stats['total_rows']}")
        print(f"   Noise rows removed: {self.stats['noise_removed']}")
        print(f"   Empty rows removed: {self.stats['empty_removed']}")
        print(f"   Valid requirements: {self.stats['valid_requirements']}")
        
        if source_type == "PDF":
            print(f"   Tables found: {self.stats['pdf_tables_found']}")
            print(f"   Rows extracted: {self.stats['pdf_rows_extracted']}")
    
    def _print_debug_info(self, ws):
        """Print first 10 rows for debugging"""
        print("\n🔍 First 10 rows for debugging:")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, 6):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    row_values.append(str(cell.value)[:20])
            print(f"   Row {row_idx}: {', '.join(row_values)}")