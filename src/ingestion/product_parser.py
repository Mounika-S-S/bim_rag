# src/ingestion/product_parser.py
"""
Enhanced Product Extractor for Layer 2 products
Handles both Excel and PDF inputs with intelligent parsing
"""
import re
import os
import pdfplumber
from openpyxl import load_workbook
from src.core.schema import create_layer_record
from src.utils.id_generator import generate_id


class ProductExtractor:
    """
    Extracts product data from Excel and PDF files
    
    Features:
    - Excel: Header detection, multiple sheets, data cleaning
    - PDF: Pattern-based extraction, text cleaning
    - Standardized output format
    """
    
    # Common product attributes to look for
    EXPECTED_ATTRIBUTES = [
        "product_name", "manufacturer", "model_number", 
        "fire_rating_hours", "compressive_strength_mpa", 
        "length_mm", "width_mm", "depth_mm", "thickness_mm",
        "weight_kg", "unit_cost_inr", "warranty_years",
        "applicable_standards", "material", "color", "finish"
    ]
    
    # PDF regex patterns for extraction
    PDF_PATTERNS = {
        "product_name": [
            r"Product Name:\s*(.*)",
            r"Product:\s*(.*)",
            r"Name:\s*(.*)"
        ],
        "manufacturer": [
            r"Manufacturer:\s*(.*)",
            r"Brand:\s*(.*)",
            r"Made by:\s*(.*)"
        ],
        "model_number": [
            r"Model Number:\s*(.*)",
            r"Model:\s*(.*)",
            r"Model No[.:]\s*(.*)"
        ],
        "fire_rating_hours": [
            r"Fire Rating\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:hours?|h)",
            r"Fire Resistance\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:min|minutes?)",
            r"EI\s*(\d+)"
        ],
        "compressive_strength_mpa": [
            r"Compressive Strength\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:MPa|N/mm2)",
            r"Strength\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:MPa)"
        ],
        "length_mm": [
            r"Length\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:mm|millimeters?)",
            r"L\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:mm)"
        ],
        "width_mm": [
            r"Width\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:mm|millimeters?)",
            r"W\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:mm)"
        ],
        "depth_mm": [
            r"Depth\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:mm|millimeters?)",
            r"D\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:mm)"
        ],
        "thickness_mm": [
            r"Thickness\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:mm|millimeters?)",
            r"T\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:mm)"
        ],
        "weight_kg": [
            r"Weight\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:kg|kilograms?)",
            r"Mass\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:kg)"
        ],
        "warranty_years": [
            r"Warranty\s*(?:Period)?\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:years?|yrs?)",
            r"Guarantee\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:years?)"
        ]
    }
    
    def __init__(self):
        self.stats = {
            "pdf_files_processed": 0,
            "excel_files_processed": 0,
            "products_extracted": 0,
            "pdf_pattern_matches": 0
        }
    
    # ============================
    # Public Methods
    # ============================
    
    def extract_from_pdf(self, file_path):
        """Extract products from PDF file"""
        self.stats["pdf_files_processed"] += 1
        return self._parse_pdf(file_path)
    
    def extract_from_excel(self, file_path, sheet_name=None):
        """Extract products from Excel file"""
        self.stats["excel_files_processed"] += 1
        return self._parse_excel(file_path, sheet_name)
    
    def extract(self, file_path, file_type=None, sheet_name=None):
        """
        Unified extraction method - auto-detects file type
        
        Args:
            file_path: Path to Excel or PDF file
            file_type: 'excel' or 'pdf' (auto-detected if None)
            sheet_name: Sheet name for Excel files
        
        Returns:
            List of product records
        """
        if file_type is None:
            ext = os.path.splitext(file_path)[1].lower()
            file_type = 'excel' if ext in ['.xlsx', '.xls', '.xlsm'] else 'pdf'
        
        print(f"\n📄 Extracting products from {file_type.upper()}: {os.path.basename(file_path)}")
        
        if file_type == 'excel':
            return self._parse_excel(file_path, sheet_name)
        else:
            return self._parse_pdf(file_path)
    
    # ============================
    # PDF Parsing (Enhanced)
    # ============================
    
    def _parse_pdf(self, file_path):
        """Enhanced PDF parsing with multiple pattern matching"""
        products = []
        
        try:
            # Extract text from PDF
            with pdfplumber.open(file_path) as pdf:
                full_text = ""
                for page_num, page in enumerate(pdf.pages, 1):
                    text = page.extract_text() or ""
                    full_text += text + "\n"
                    
                    # Also try to extract tables
                    tables = page.extract_tables()
                    for table in tables:
                        table_products = self._parse_pdf_table(table, page_num)
                        products.extend(table_products)
            
            if not full_text.strip():
                print("⚠️ No text extracted from PDF")
                return products
            
            # Clean text
            full_text = self._clean_text(full_text)
            
            # Try to find product sections
            product_sections = self._split_into_products(full_text)
            
            if product_sections:
                # Multiple products found
                for section in product_sections:
                    product_data = self._extract_from_text(section)
                    if product_data:
                        record = self._create_product_record(product_data, file_path)
                        products.append(record)
                        self.stats["products_extracted"] += 1
            else:
                # Single product or no clear separation
                product_data = self._extract_from_text(full_text)
                if product_data:
                    record = self._create_product_record(product_data, file_path)
                    products.append(record)
                    self.stats["products_extracted"] += 1
            
            print(f"✅ Extracted {len(products)} products from PDF")
            return products
            
        except Exception as e:
            print(f"❌ PDF parsing error: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def _parse_pdf_table(self, table, page_num):
        """Extract products from PDF tables"""
        products = []
        
        if not table or len(table) < 2:
            return products
        
        # Assume first row is headers
        headers = [str(cell).strip().lower() if cell else f"col_{i}" 
                  for i, cell in enumerate(table[0])]
        
        # Process data rows
        for row in table[1:]:
            if not row or all(cell is None for cell in row):
                continue
            
            properties = {}
            for i, cell in enumerate(row):
                if i < len(headers) and cell:
                    key = self._normalize_header(headers[i])
                    value = str(cell).strip()
                    if value:
                        properties[key] = self._convert_value(value)
            
            if properties:
                record = self._create_product_record(
                    properties, 
                    f"PDF Table (Page {page_num})"
                )
                products.append(record)
                self.stats["products_extracted"] += 1
                self.stats["pdf_pattern_matches"] += 1
        
        return products
    
    def _split_into_products(self, text):
        """Split text into individual product sections"""
        # Look for common product separators
        separators = [
            r"Product \d+",
            r"Item \d+",
            r"---+\s*\n",
            r"===+\s*\n",
            r"\n\s*\n\s*[A-Z][a-z]+"  # New line with capitalized word
        ]
        
        for separator in separators:
            sections = re.split(separator, text)
            if len(sections) > 1:
                return sections
        
        return [text]  # Return as single section
    
    def _extract_from_text(self, text):
        """Extract product data from text using patterns"""
        product_data = {}
        
        for key, patterns in self.PDF_PATTERNS.items():
            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    value = match.group(1).strip()
                    # Convert to appropriate type
                    if any(x in key for x in ['rating', 'strength', 'length', 'width', 'depth', 'thickness', 'weight', 'cost']):
                        try:
                            # Extract number from string (e.g., "120 min" -> 120)
                            num_match = re.search(r'(\d+(?:\.\d+)?)', value)
                            if num_match:
                                value = float(num_match.group(1))
                        except:
                            pass
                    
                    product_data[key] = value
                    self.stats["pdf_pattern_matches"] += 1
                    break
        
        return product_data
    
    # ============================
    # Excel Parsing (Enhanced)
    # ============================
    
    def _parse_excel(self, file_path, sheet_name=None):
        """Enhanced Excel parsing with header detection"""
        products = []
        
        try:
            wb = load_workbook(file_path, data_only=True)
            
            # Select worksheet
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"📑 Using sheet: {sheet_name}")
            else:
                ws = wb.active
                print(f"📑 Using active sheet: {ws.title}")
            
            # Find header row (search first 10 rows)
            header_row_idx, headers = self._find_header_row(ws)
            
            if not headers:
                print("⚠️ Could not find header row, using first row")
                headers = [cell.value for cell in ws[1] if cell.value]
                header_row_idx = 1
            
            # Normalize headers
            normalized_headers = [self._normalize_header(h) for h in headers]
            
            # Process data rows
            empty_count = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), 
                                          start=header_row_idx + 1):
                
                if not row or all(cell is None for cell in row):
                    empty_count += 1
                    if empty_count > 5:  # Stop after 5 empty rows
                        break
                    continue
                
                empty_count = 0
                
                # Create properties dictionary
                properties = {}
                for i, value in enumerate(row):
                    if i < len(normalized_headers) and value is not None:
                        key = normalized_headers[i]
                        properties[key] = self._convert_value(value)
                
                if properties:
                    record = self._create_product_record(properties, file_path)
                    products.append(record)
                    self.stats["products_extracted"] += 1
            
            print(f"✅ Extracted {len(products)} products from Excel")
            return products
            
        except Exception as e:
            print(f"❌ Excel parsing error: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def _find_header_row(self, ws):
        """Find the row containing headers by looking for common patterns"""
        expected_keywords = ['product', 'name', 'manufacturer', 'model', 'description']
        
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row = [cell.value for cell in ws[row_idx] if cell.value]
            if not row:
                continue
            
            row_text = ' '.join([str(c).lower() for c in row if c])
            
            # Check if row contains expected keywords
            matches = sum(1 for kw in expected_keywords if kw in row_text)
            if matches >= 2:  # At least 2 matches
                return row_idx, [cell.value for cell in ws[row_idx]]
        
        return 1, [cell.value for cell in ws[1] if cell.value]
    
    def _normalize_header(self, header):
        """Convert Excel header to standard attribute name"""
        if not header:
            return f"attribute_{hash(str(header)) % 1000}"
        
        # Convert to string and lowercase
        header_str = str(header).lower().strip()
        
        # Replace spaces and special chars with underscore
        header_str = re.sub(r'[^\w\s]', '', header_str)
        header_str = re.sub(r'\s+', '_', header_str)
        
        # Common mappings
        mappings = {
            'product': 'product_name',
            'prod_name': 'product_name',
            'name': 'product_name',
            'manufacturer': 'manufacturer',
            'mfr': 'manufacturer',
            'brand': 'manufacturer',
            'model': 'model_number',
            'model_no': 'model_number',
            'fire': 'fire_rating_hours',
            'fire_rating': 'fire_rating_hours',
            'strength': 'compressive_strength_mpa',
            'compressive_strength': 'compressive_strength_mpa',
            'length': 'length_mm',
            'width': 'width_mm',
            'depth': 'depth_mm',
            'thickness': 'thickness_mm',
            'weight': 'weight_kg',
            'cost': 'unit_cost_inr',
            'price': 'unit_cost_inr',
            'warranty': 'warranty_years'
        }
        
        # Check if header matches any mapping
        for key, mapped in mappings.items():
            if key in header_str:
                return mapped
        
        return header_str
    
    def _convert_value(self, value):
        """Convert string values to appropriate types"""
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return value
        
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            
            # Try to convert to number
            try:
                # Remove currency symbols and commas
                clean = re.sub(r'[₹$,]', '', value)
                return float(clean)
            except:
                return value
        
        return value
    
    # ============================
    # Record Creation
    # ============================
    
    def _create_product_record(self, properties, source):
        """Create standardized product record"""
        
        # Ensure required fields
        if 'product_name' not in properties and 'name' in properties:
            properties['product_name'] = properties.pop('name')
        
        # Add metadata
        properties['source_document'] = os.path.basename(source) if isinstance(source, str) else source
        properties['extraction_timestamp'] = str(__import__('datetime').datetime.now())
        
        # Generate ID
        if 'product_name' in properties and 'manufacturer' in properties:
            # Create readable ID
            name_part = re.sub(r'[^a-zA-Z0-9]', '', properties['product_name'])[:10]
            mfr_part = re.sub(r'[^a-zA-Z0-9]', '', properties['manufacturer'])[:5]
            record_id = f"PROD_{mfr_part}_{name_part}"
        else:
            record_id = generate_id("L2")
        
        return create_layer_record(
            record_id=record_id,
            entity_type="Product",
            layer="L2",
            category="Technical",
            properties=properties
        )
    
    # ============================
    # Text Cleaning
    # ============================
    
    def _clean_text(self, text):
        """Clean extracted text"""
        # Remove extra whitespace
        text = re.sub(r'\s+', ' ', text)
        
        # Remove bullet points and special characters
        text = re.sub(r'[•●▪➢➤]', '', text)
        
        # Fix common OCR issues
        text = text.replace('|', '')
        text = text.replace('®', '')
        text = text.replace('™', '')
        
        return text.strip()
    
    def get_stats(self):
        """Return extraction statistics"""
        return self.stats